from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = PROJECT_ROOT / "data" / "processed" / "ionic_liquid_6_properties_values_errors.xlsx"
DEFAULT_COMPOUNDS = PROJECT_ROOT / "data" / "ILThermoPy-main" / "src" / "ilthermopy" / "data" / "compounds.csv"
DEFAULT_MISSING = PROJECT_ROOT / "data" / "processed" / "ionic_liquid_missing_smiles_names.csv"


def normalize_name(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\xa0", " ")
    return re.sub(r"\s+", " ", text.strip().lower())


def load_name_to_smiles(path: Path) -> dict[str, str]:
    mapping: dict[str, str] = {}
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"name", "smiles"}
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise ValueError(f"Missing required columns in {path}: {', '.join(sorted(missing))}")
        for row in reader:
            name = normalize_name(row["name"])
            smiles = (row["smiles"] or "").strip()
            if name and smiles:
                mapping[name] = smiles
    return mapping


def header_index(headers: list[object], name: str) -> int:
    try:
        return headers.index(name) + 1
    except ValueError as exc:
        available = ", ".join(str(h) for h in headers)
        raise ValueError(f"Missing required column {name!r}. Available columns: {available}") from exc


def copy_column_style(ws, source_col: int, target_col: int) -> None:
    ws.column_dimensions[ws.cell(1, target_col).column_letter].width = max(
        ws.column_dimensions[ws.cell(1, source_col).column_letter].width or 12,
        32,
    )
    for row in range(1, ws.max_row + 1):
        source = ws.cell(row, source_col)
        target = ws.cell(row, target_col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def add_il_smiles(
    workbook_path: Path,
    compounds_path: Path,
    sheet_name: str,
    missing_report_path: Path | None,
) -> tuple[int, int, int]:
    name_to_smiles = load_name_to_smiles(compounds_path)
    wb = load_workbook(workbook_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    il_col = header_index(headers, "IL_Name")

    if "IL_SMILES" in headers:
        smiles_col = header_index(headers, "IL_SMILES")
    else:
        smiles_col = il_col + 1
        ws.insert_cols(smiles_col)
        copy_column_style(ws, il_col, smiles_col)
        ws.cell(1, smiles_col).value = "IL_SMILES"

    rows = ws.max_row - 1
    matched = 0
    missing_counter: Counter[str] = Counter()
    for row_idx in range(2, ws.max_row + 1):
        il_name = ws.cell(row_idx, il_col).value
        smiles = name_to_smiles.get(normalize_name(il_name), "")
        ws.cell(row_idx, smiles_col).value = smiles
        if smiles:
            matched += 1
        else:
            missing_counter[str(il_name or "")] += 1

    if missing_report_path:
        missing_report_path.parent.mkdir(parents=True, exist_ok=True)
        with missing_report_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["IL_Name", "row_count"])
            for name, count in sorted(missing_counter.items(), key=lambda item: (-item[1], item[0])):
                writer.writerow([name, count])

    wb.save(workbook_path)
    return rows, matched, len(missing_counter)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Insert IL_SMILES after IL_Name using the local ILThermo compounds table.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--compounds", type=Path, default=DEFAULT_COMPOUNDS)
    parser.add_argument("--sheet", default="Merged")
    parser.add_argument("--missing-report", type=Path, default=DEFAULT_MISSING)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows, matched, missing = add_il_smiles(
        workbook_path=args.workbook,
        compounds_path=args.compounds,
        sheet_name=args.sheet,
        missing_report_path=args.missing_report,
    )
    print(f"workbook: {args.workbook}")
    print(f"rows: {rows}")
    print(f"rows with IL_SMILES: {matched}")
    print(f"rows without IL_SMILES: {rows - matched}")
    print(f"unique unresolved IL names: {missing}")
    if args.missing_report:
        print(f"missing report: {args.missing_report}")


if __name__ == "__main__":
    main()
