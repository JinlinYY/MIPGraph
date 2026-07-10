from __future__ import annotations

import argparse
import bisect
import csv
import json
import math
from collections import defaultdict
from copy import copy
from pathlib import Path
from statistics import median
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font


PROPERTIES = (
    "Density",
    "ElectricalConductivity",
    "HeatCapacity",
    "SurfaceTension",
    "ThermalConductivity",
    "Viscosity",
)
LOG_PROPERTIES = {"ElectricalConductivity", "Viscosity"}


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip()) or (
        isinstance(value, float) and math.isnan(value)
    )


def as_float(value: Any) -> float | None:
    if is_blank(value):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def interpolate(left: float, right: float, alpha: float, log_space: bool) -> float:
    if log_space:
        return math.exp((1.0 - alpha) * math.log(left) + alpha * math.log(right))
    return (1.0 - alpha) * left + alpha * right


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fill missing properties by same-IL, same-pressure interpolation without extrapolation."
    )
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--summary", type=Path, required=True)
    parser.add_argument("--sheet", default="Merged")
    parser.add_argument("--properties", default=",".join(PROPERTIES))
    parser.add_argument("--max-temperature-gap", type=float, default=40.0)
    parser.add_argument("--pressure-round-decimals", type=int, default=1)
    parser.add_argument("--missing-pressure-kpa", type=float, default=101.325)
    parser.add_argument("--max-replicate-relative-spread", type=float, default=0.15)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    properties = tuple(item.strip() for item in args.properties.split(",") if item.strip())
    unknown = sorted(set(properties) - set(PROPERTIES))
    if unknown:
        raise ValueError(f"Unknown properties: {unknown}")
    if args.max_temperature_gap <= 0:
        raise ValueError("--max-temperature-gap must be positive")

    wb = load_workbook(args.input)
    ws = wb[args.sheet]
    columns = {str(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}
    for name in ("IL_Name", "Temperature_K", "Pressure_kPa"):
        if name not in columns:
            raise ValueError(f"Missing required column: {name}")

    source_columns: dict[str, int] = {}
    for prop in properties:
        value_column = f"{prop}_ActualValue"
        error_column = f"{prop}_ErrorValue"
        if value_column not in columns or error_column not in columns:
            raise ValueError(f"Missing required property columns for {prop}")
        source_column = f"{prop}_ValueSource"
        if source_column not in columns:
            col = ws.max_column + 1
            ws.cell(1, col).value = source_column
            ws.cell(1, col)._style = copy(ws.cell(1, col - 1)._style)
            columns[source_column] = col
        source_columns[prop] = columns[source_column]

    il_col = columns["IL_Name"]
    temp_col = columns["Temperature_K"]
    pressure_col = columns["Pressure_kPa"]
    report_rows: list[dict[str, Any]] = []
    summary: dict[str, dict[str, int]] = {}

    for prop in properties:
        value_col = columns[f"{prop}_ActualValue"]
        error_col = columns[f"{prop}_ErrorValue"]
        source_col = source_columns[prop]
        observations: dict[tuple[str, float], dict[float, list[tuple[float, float | None, int]]]] = defaultdict(
            lambda: defaultdict(list)
        )
        targets: list[tuple[int, str, float, float]] = []

        for row_idx in range(2, ws.max_row + 1):
            il_name = normalize_text(ws.cell(row_idx, il_col).value)
            temperature = as_float(ws.cell(row_idx, temp_col).value)
            pressure = as_float(ws.cell(row_idx, pressure_col).value)
            pressure = args.missing_pressure_kpa if pressure is None else pressure
            if not il_name or temperature is None:
                continue
            pressure_key = round(pressure, args.pressure_round_decimals)
            value = as_float(ws.cell(row_idx, value_col).value)
            error = as_float(ws.cell(row_idx, error_col).value)
            if value is not None and value > 0:
                observations[(il_name, pressure_key)][temperature].append((value, error, row_idx))
            else:
                targets.append((row_idx, il_name, temperature, pressure_key))

        curves: dict[tuple[str, float], list[dict[str, Any]]] = {}
        conflicting_knots = 0
        for key, by_temperature in observations.items():
            knots = []
            for temperature, records in by_temperature.items():
                values = [record[0] for record in records]
                center = float(median(values))
                relative_spread = (max(values) - min(values)) / max(abs(center), 1e-12)
                if relative_spread > args.max_replicate_relative_spread:
                    conflicting_knots += 1
                    continue
                errors = [record[1] for record in records if record[1] is not None and record[1] >= 0]
                knots.append(
                    {
                        "temperature": float(temperature),
                        "value": center,
                        "error": float(median(errors)) if errors else None,
                        "source_rows": ";".join(str(record[2]) for record in records),
                    }
                )
            if knots:
                curves[key] = sorted(knots, key=lambda item: item["temperature"])

        counts = defaultdict(int)
        for row_idx, il_name, temperature, pressure_key in targets:
            knots = curves.get((il_name, pressure_key), [])
            if not knots:
                counts["no_curve"] += 1
                continue
            temperatures = [item["temperature"] for item in knots]
            pos = bisect.bisect_left(temperatures, temperature)
            exact = pos < len(knots) and abs(temperatures[pos] - temperature) <= 1e-8
            if exact:
                left = right = knots[pos]
                alpha = 0.0
                method = "exact_condition_copy"
                value = left["value"]
                error = left["error"]
            else:
                if pos == 0 or pos == len(knots):
                    counts["outside_observed_range"] += 1
                    continue
                left, right = knots[pos - 1], knots[pos]
                gap = right["temperature"] - left["temperature"]
                if gap <= 0 or gap > args.max_temperature_gap:
                    counts["temperature_gap_too_large"] += 1
                    continue
                alpha = (temperature - left["temperature"]) / gap
                value = interpolate(left["value"], right["value"], alpha, prop in LOG_PROPERTIES)
                if left["error"] is not None and right["error"] is not None:
                    error = (1.0 - alpha) * left["error"] + alpha * right["error"]
                else:
                    error = None
                method = "temperature_log_interpolation" if prop in LOG_PROPERTIES else "temperature_linear_interpolation"

            if not math.isfinite(value) or value <= 0:
                counts["invalid_interpolated_value"] += 1
                continue
            ws.cell(row_idx, value_col).value = value
            ws.cell(row_idx, source_col).value = method
            if is_blank(ws.cell(row_idx, error_col).value) and error is not None:
                ws.cell(row_idx, error_col).value = error
            counts[method] += 1
            report_rows.append(
                {
                    "row": row_idx,
                    "IL_Name": il_name,
                    "property": prop,
                    "Temperature_K": temperature,
                    "Pressure_kPa_normalized": pressure_key,
                    "filled_value": value,
                    "filled_error": error if error is not None else "",
                    "method": method,
                    "alpha": alpha,
                    "left_temperature": left["temperature"],
                    "left_value": left["value"],
                    "left_source_rows": left["source_rows"],
                    "right_temperature": right["temperature"],
                    "right_value": right["value"],
                    "right_source_rows": right["source_rows"],
                }
            )
        counts["conflicting_source_knots"] = conflicting_knots
        counts["filled_total"] = counts["exact_condition_copy"] + counts["temperature_linear_interpolation"] + counts[
            "temperature_log_interpolation"
        ]
        summary[prop] = dict(counts)

    for sheet_name in ("InterpolationProvenance", "InterpolationSummary"):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    provenance_ws = wb.create_sheet("InterpolationProvenance")
    fields = list(report_rows[0].keys()) if report_rows else ["row", "property", "method"]
    provenance_ws.append(fields)
    for row in report_rows:
        provenance_ws.append([row.get(field, "") for field in fields])
    for cell in provenance_ws[1]:
        cell.font = Font(name="Arial", bold=True)

    summary_ws = wb.create_sheet("InterpolationSummary")
    summary_fields = [
        "property",
        "filled_total",
        "exact_condition_copy",
        "temperature_linear_interpolation",
        "temperature_log_interpolation",
        "no_curve",
        "outside_observed_range",
        "temperature_gap_too_large",
        "conflicting_source_knots",
    ]
    summary_ws.append(summary_fields)
    for prop in properties:
        summary_ws.append([prop] + [summary[prop].get(field, 0) for field in summary_fields[1:]])
    for cell in summary_ws[1]:
        cell.font = Font(name="Arial", bold=True)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    with args.report.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(report_rows)
    args.summary.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    print(f"output: {args.output}")
    print(f"report: {args.report}")
    print(f"summary: {args.summary}")


if __name__ == "__main__":
    main()
