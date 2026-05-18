from __future__ import annotations

import argparse
import csv
import json
import math
import sys
import time
from collections import Counter, OrderedDict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from tqdm import tqdm
except ImportError:  # pragma: no cover - fallback for minimal environments.
    tqdm = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
ILTHERMOPY_SRC = PROJECT_ROOT / "data" / "ILThermoPy-main" / "src"
DEFAULT_INPUT = PROJECT_ROOT / "data" / "processed" / "ionic_liquid_6_properties_values_errors.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "processed" / "ionic_liquid_6_properties_values_errors_ilthermo_strict.xlsx"
DEFAULT_CACHE = PROJECT_ROOT / "data" / "cache" / "ilthermopy_strict_property_cache.json"
DEFAULT_REPORT = PROJECT_ROOT / "data" / "processed" / "ilthermopy_strict_property_fill_report.csv"


@dataclass(frozen=True)
class PropertySpec:
    name: str
    prop_key: str
    value_column: str
    error_column: str
    aliases: tuple[str, ...]


PROPERTY_SPECS = (
    PropertySpec(
        name="Density",
        prop_key="JkYu",
        value_column="Density_ActualValue",
        error_column="Density_ErrorValue",
        aliases=("density",),
    ),
    PropertySpec(
        name="ElectricalConductivity",
        prop_key="Ylwl",
        value_column="ElectricalConductivity_ActualValue",
        error_column="ElectricalConductivity_ErrorValue",
        aliases=("electrical conductivity",),
    ),
    PropertySpec(
        name="HeatCapacity",
        prop_key="IZSt",
        value_column="HeatCapacity_ActualValue",
        error_column="HeatCapacity_ErrorValue",
        aliases=("heat capacity at constant pressure", "heat capacity"),
    ),
    PropertySpec(
        name="SurfaceTension",
        prop_key="ETUw",
        value_column="SurfaceTension_ActualValue",
        error_column="SurfaceTension_ErrorValue",
        aliases=("surface tension liquid-gas", "surface tension"),
    ),
    PropertySpec(
        name="ThermalConductivity",
        prop_key="pAFI",
        value_column="ThermalConductivity_ActualValue",
        error_column="ThermalConductivity_ErrorValue",
        aliases=("thermal conductivity",),
    ),
    PropertySpec(
        name="Viscosity",
        prop_key="PusA",
        value_column="Viscosity_ActualValue",
        error_column="Viscosity_ErrorValue",
        aliases=("viscosity",),
    ),
)

PROPERTY_BY_NAME = {spec.name.lower(): spec for spec in PROPERTY_SPECS}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").strip().split())


def normalize_key(value: Any) -> str:
    return normalize_text(value).lower()


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, float):
        return math.isnan(value)
    return False


def number_key(value: Any) -> str:
    if is_blank(value):
        return ""
    try:
        dec = Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return normalize_text(value)
    if dec.is_zero():
        return "0"
    return format(dec.normalize(), "f")


def json_number(value: Any) -> float | None:
    if is_blank(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def cache_lookup_key(il_name: str, spec: PropertySpec) -> str:
    return f"{spec.prop_key}:{normalize_key(il_name)}"


def load_cache(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"records": {}}
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if "records" not in data:
        return {"records": data}
    return data


def save_cache(path: Path, cache: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(cache, handle, ensure_ascii=False, indent=2)


def import_ilthermopy(request_timeout: float):
    sys.path.insert(0, str(ILTHERMOPY_SRC))
    import ilthermopy as ilt
    import ilthermopy.requests as ilt_requests

    original_get = ilt_requests._requests.get

    def get_with_timeout(*args, **kwargs):
        kwargs.setdefault("timeout", request_timeout)
        return original_get(*args, **kwargs)

    ilt_requests._requests.get = get_with_timeout

    return ilt


def find_header(headers: dict[str, str], predicate) -> str | None:
    for column, title in headers.items():
        if predicate(normalize_key(title)):
            return column
    return None


def header_has_alias(header: str, spec: PropertySpec) -> bool:
    return any(alias in header for alias in spec.aliases)


def extract_entry_rows(entry: Any, il_name: str, spec: PropertySpec) -> list[dict[str, Any]]:
    if len(entry.components) != 1:
        return []
    if normalize_key(entry.components[0].name) != normalize_key(il_name):
        return []

    temp_col = find_header(entry.header, lambda h: h.startswith("temperature"))
    pressure_col = find_header(entry.header, lambda h: h.startswith("pressure"))
    value_col = find_header(
        entry.header,
        lambda h: (
            header_has_alias(h, spec)
            and not h.startswith("error of")
            and not h.startswith("temperature")
            and not h.startswith("pressure")
        ),
    )
    error_col = find_header(
        entry.header,
        lambda h: h.startswith("error of") and header_has_alias(h, spec),
    )
    if temp_col is None or value_col is None:
        return []

    rows: list[dict[str, Any]] = []
    for _, data_row in entry.data.iterrows():
        rows.append(
            {
                "temperature_key": number_key(data_row[temp_col]),
                "pressure_key": number_key(data_row[pressure_col]) if pressure_col else "",
                "actual": json_number(data_row[value_col]),
                "error": json_number(data_row[error_col]) if error_col else None,
                "entry_id": entry.id,
            }
        )
    return rows


def query_ilthermo(ilt: Any, il_name: str, spec: PropertySpec) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    search = ilt.Search(compound=il_name, n_compounds=1, prop_key=spec.prop_key)
    for _, result in search.iterrows():
        if normalize_key(result.get("cmp1")) != normalize_key(il_name):
            continue
        entry = ilt.GetEntry(str(result["id"]))
        rows.extend(extract_entry_rows(entry, il_name, spec))
    return {"status": "ok", "rows": rows, "error": None}


def resolve_candidates(candidates: list[dict[str, Any]]) -> tuple[str, dict[str, Any] | None]:
    if not candidates:
        return "no_match", None

    actual_keys = {number_key(candidate.get("actual")) for candidate in candidates}
    error_keys = {number_key(candidate.get("error")) for candidate in candidates}
    if len(actual_keys) > 1 or len(error_keys) > 1:
        return "ambiguous", None

    resolved = dict(candidates[0])
    resolved["entry_id"] = ";".join(sorted({str(candidate.get("entry_id", "")) for candidate in candidates}))
    return "matched", resolved


def header_to_col(ws) -> dict[str, int]:
    return {str(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1)}


def collect_needed_queries(
    ws,
    columns: dict[str, int],
    property_specs: tuple[PropertySpec, ...],
) -> OrderedDict[str, tuple[str, PropertySpec]]:
    il_col = columns["IL_Name"]
    needed: OrderedDict[str, tuple[str, PropertySpec]] = OrderedDict()
    for row_idx in range(2, ws.max_row + 1):
        il_name = normalize_text(ws.cell(row_idx, il_col).value)
        if not il_name:
            continue
        for spec in property_specs:
            value_col = columns.get(spec.value_column)
            error_col = columns.get(spec.error_column)
            if value_col is None or error_col is None:
                continue
            if is_blank(ws.cell(row_idx, value_col).value) or is_blank(ws.cell(row_idx, error_col).value):
                needed.setdefault(cache_lookup_key(il_name, spec), (il_name, spec))
    return needed


def build_cache(
    needed: OrderedDict[str, tuple[str, PropertySpec]],
    cache: dict[str, Any],
    cache_path: Path,
    retry_failed: bool,
    limit: int | None,
    sleep_seconds: float,
    checkpoint_every: int,
    progress_every: int,
    request_timeout: float,
    use_tqdm: bool,
) -> Counter[str]:
    ilt = import_ilthermopy(request_timeout=request_timeout)
    stats: Counter[str] = Counter()
    records = cache.setdefault("records", {})
    queried = 0
    pending_all = [
        (key, il_name, spec)
        for key, (il_name, spec) in needed.items()
        if not (records.get(key) and (records[key].get("status") == "ok" or not retry_failed))
    ]
    pending_items = pending_all
    if limit is not None:
        pending_items = pending_all[:limit]
        stats["limited"] += max(0, len(pending_all) - len(pending_items))

    for key, (il_name, spec) in needed.items():
        existing = records.get(key)
        if existing and (existing.get("status") == "ok" or not retry_failed):
            stats["cache_hits"] += 1

    iterator = pending_items
    progress_bar = None
    if use_tqdm and tqdm is not None:
        progress_bar = tqdm(pending_items, desc="ILThermo queries", unit="query")
        iterator = progress_bar

    for key, il_name, spec in iterator:
        queried += 1
        stats["queries"] += 1
        if progress_bar is not None:
            progress_bar.set_postfix_str(f"{spec.name}: {il_name[:45]}")
        elif progress_every and queried % progress_every == 1:
            print(f"query {queried}: {spec.name} | {il_name}", flush=True)
        try:
            records[key] = query_ilthermo(ilt, il_name, spec)
            stats["query_ok"] += 1
        except Exception as exc:  # noqa: BLE001 - persist query errors for resumable runs.
            records[key] = {
                "status": "error",
                "rows": [],
                "error": f"{type(exc).__name__}: {exc}",
            }
            stats["query_error"] += 1

        if checkpoint_every and queried % checkpoint_every == 0:
            save_cache(cache_path, cache)
        if progress_bar is None and progress_every and queried % progress_every == 0:
            print(f"queried {queried} new IL/property pairs; cache saved={bool(checkpoint_every)}", flush=True)
        if sleep_seconds:
            time.sleep(sleep_seconds)

    return stats


def fill_properties(
    ws,
    columns: dict[str, int],
    cache: dict[str, Any],
    property_specs: tuple[PropertySpec, ...],
    use_tqdm: bool,
) -> tuple[Counter[str], list[dict[str, Any]]]:
    il_col = columns["IL_Name"]
    temp_col = columns["Temperature_K"]
    pressure_col = columns["Pressure_kPa"]
    records = cache.get("records", {})
    stats: Counter[str] = Counter()
    report_rows: list[dict[str, Any]] = []

    row_iter = range(2, ws.max_row + 1)
    if use_tqdm and tqdm is not None:
        row_iter = tqdm(row_iter, desc="Filling workbook rows", unit="row")

    for row_idx in row_iter:
        il_name = normalize_text(ws.cell(row_idx, il_col).value)
        if not il_name:
            continue
        temp_key = number_key(ws.cell(row_idx, temp_col).value)
        pressure_key = number_key(ws.cell(row_idx, pressure_col).value)

        for spec in property_specs:
            value_col = columns.get(spec.value_column)
            error_col = columns.get(spec.error_column)
            if value_col is None or error_col is None:
                continue
            value_cell = ws.cell(row_idx, value_col)
            error_cell = ws.cell(row_idx, error_col)
            needs_value = is_blank(value_cell.value)
            needs_error = is_blank(error_cell.value)
            if not needs_value and not needs_error:
                continue

            stats["missing_cells_seen"] += int(needs_value) + int(needs_error)
            record = records.get(cache_lookup_key(il_name, spec), {})
            if record.get("status") != "ok":
                status = "query_error" if record else "not_queried"
                stats[status] += 1
                report_rows.append(
                    {
                        "row": row_idx,
                        "IL_Name": il_name,
                        "property": spec.name,
                        "Temperature_K": ws.cell(row_idx, temp_col).value,
                        "Pressure_kPa": ws.cell(row_idx, pressure_col).value,
                        "status": status,
                        "filled_actual": "",
                        "filled_error": "",
                        "source_entry_ids": "",
                        "note": record.get("error", ""),
                    }
                )
                continue

            candidates = [
                row
                for row in record.get("rows", [])
                if row.get("temperature_key") == temp_key and row.get("pressure_key") == pressure_key
            ]
            status, resolved = resolve_candidates(candidates)
            stats[status] += 1
            filled_actual = ""
            filled_error = ""
            source_ids = ""
            note = ""

            if status == "matched" and resolved:
                source_ids = resolved.get("entry_id", "")
                if needs_value and resolved.get("actual") is not None:
                    value_cell.value = resolved["actual"]
                    filled_actual = resolved["actual"]
                    stats["actual_values_filled"] += 1
                if needs_error and resolved.get("error") is not None:
                    error_cell.value = resolved["error"]
                    filled_error = resolved["error"]
                    stats["error_values_filled"] += 1
                if filled_actual == "" and filled_error == "":
                    note = "candidate matched, but requested cells had no available value/error"
            elif status == "ambiguous":
                note = "multiple strict matches had conflicting values"

            report_rows.append(
                {
                    "row": row_idx,
                    "IL_Name": il_name,
                    "property": spec.name,
                    "Temperature_K": ws.cell(row_idx, temp_col).value,
                    "Pressure_kPa": ws.cell(row_idx, pressure_col).value,
                    "status": status,
                    "filled_actual": filled_actual,
                    "filled_error": filled_error,
                    "source_entry_ids": source_ids,
                    "note": note,
                }
            )

    return stats, report_rows


def write_report(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "row",
        "IL_Name",
        "property",
        "Temperature_K",
        "Pressure_kPa",
        "status",
        "filled_actual",
        "filled_error",
        "source_entry_ids",
        "note",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def parse_property_specs(value: str | None) -> tuple[PropertySpec, ...]:
    if not value:
        return PROPERTY_SPECS
    specs: list[PropertySpec] = []
    for item in value.split(","):
        name = item.strip().lower()
        if not name:
            continue
        spec = PROPERTY_BY_NAME.get(name)
        if spec is None:
            available = ", ".join(spec.name for spec in PROPERTY_SPECS)
            raise ValueError(f"Unknown property {item!r}. Available: {available}")
        specs.append(spec)
    return tuple(dict.fromkeys(specs))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Strictly fill missing property values from ILThermoPy by IL name, property, temperature, and pressure."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--in-place", action="store_true", help="Overwrite --input instead of writing --output.")
    parser.add_argument("--sheet", default="Merged")
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--retry-failed", action="store_true")
    parser.add_argument("--limit", type=int, default=None, help="Maximum number of new ILThermoPy property queries.")
    parser.add_argument(
        "--properties",
        default=None,
        help="Comma-separated property names to fill. Available: Density, ElectricalConductivity, HeatCapacity, SurfaceTension, ThermalConductivity, Viscosity.",
    )
    parser.add_argument("--sleep-seconds", type=float, default=0.2)
    parser.add_argument("--checkpoint-every", type=int, default=25)
    parser.add_argument("--progress-every", type=int, default=25)
    parser.add_argument("--no-tqdm", action="store_true", help="Disable tqdm progress bars and use text progress only.")
    parser.add_argument(
        "--request-timeout",
        type=float,
        default=20.0,
        help="Seconds to wait for each ILThermo HTTP request before recording an error and continuing.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = args.input if args.in_place else args.output

    wb = load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active
    columns = header_to_col(ws)
    required = {"IL_Name", "Temperature_K", "Pressure_kPa"}
    missing_required = required.difference(columns)
    if missing_required:
        raise ValueError(f"Missing required workbook columns: {', '.join(sorted(missing_required))}")

    property_specs = parse_property_specs(args.properties)
    needed = collect_needed_queries(ws, columns, property_specs)
    cache = load_cache(args.cache)
    cache_stats = build_cache(
        needed=needed,
        cache=cache,
        cache_path=args.cache,
        retry_failed=args.retry_failed,
        limit=args.limit,
        sleep_seconds=args.sleep_seconds,
        checkpoint_every=args.checkpoint_every,
        progress_every=args.progress_every,
        request_timeout=args.request_timeout,
        use_tqdm=not args.no_tqdm,
    )
    save_cache(args.cache, cache)

    fill_stats, report_rows = fill_properties(ws, columns, cache, property_specs, use_tqdm=not args.no_tqdm)
    write_report(args.report, report_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"input: {args.input}")
    print(f"output: {output_path}")
    print(f"properties: {', '.join(spec.name for spec in property_specs)}")
    print(f"unique IL/property queries needed: {len(needed)}")
    print(f"cache stats: {dict(cache_stats)}")
    print(f"fill stats: {dict(fill_stats)}")
    print(f"cache: {args.cache}")
    print(f"report: {args.report}")


if __name__ == "__main__":
    main()
